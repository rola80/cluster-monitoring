"""판정 결과를 xlsx 총괄표로 출력."""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD = PatternFill("solid", fgColor="305496")
COLORS = {"적합": "C6EFCE", "부적합": "FFC7CE", "확인필요": "FFEB9C", "해당없음": "E7E6E6"}


def _style(ws, ncol, header_row=1):
    for c in range(1, ncol + 1):
        cell = ws.cell(header_row, c)
        cell.font = Font(name="맑은 고딕", bold=True, color="FFFFFF")
        cell.fill = HEAD
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    for row in ws.iter_rows(min_row=header_row + 1, max_col=ncol):
        for cell in row:
            cell.font = Font(name="맑은 고딕")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = BORDER


def build_report(company_name, record, judgment, out_path):
    wb = Workbook()

    # 1) 종합판정
    ws = wb.active
    ws.title = "종합판정"
    ws.append(["기업명", company_name])
    ws.append(["종합판정", judgment["overall"]])
    ws.append(["신청일", str(record.get("apply_date") or "")])
    ws.append(["제출 PDF 수", record.get("n_pdfs")])
    ws.append(["가점(잠정 합계)", judgment["bonus_total"]])
    ws["B2"].fill = PatternFill("solid", fgColor=COLORS.get(judgment["overall"], "FFFFFF"))
    for r in range(1, 6):
        ws.cell(r, 1).font = Font(name="맑은 고딕", bold=True)
        for c in (1, 2):
            ws.cell(r, c).border = BORDER
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 40

    # 2) 판단기준별 결과
    ws2 = wb.create_sheet("판단기준별 결과")
    ws2.append(["구분", "ID", "판단기준", "판정", "근거/세부", "비고(evidence)", "근거조항(규정·RAG)"])
    for r in judgment["results"]:
        ws2.append([r["section"], r["id"], r["name"], r["status"], r["detail"],
                    r["evidence"], r.get("basis", "")])
    _style(ws2, 7)
    for r in range(2, ws2.max_row + 1):
        st = ws2.cell(r, 4).value
        ws2.cell(r, 4).fill = PatternFill("solid", fgColor=COLORS.get(st, "FFFFFF"))
    for col, w in zip("ABCDEFG", [20, 6, 22, 10, 45, 34, 30]):
        ws2.column_dimensions[col].width = w

    # 3) 가점 증빙
    ws3 = wb.create_sheet("가점 증빙")
    ws3.append(["ID", "가점항목", "배점", "유형", "건수", "부여(O/X)", "잠정점수", "근거파일", "비고"])
    for b in judgment["bonus"]:
        ws3.append([b["id"], b["항목"], b["배점"], b["유형"], b["건수"],
                    b["부여"], b["잠정점수"], b.get("근거파일", ""), b["비고"]])
    ws3.append(["", "합계(최대 5점)", "", "", "", "", judgment["bonus_total"], "", ""])
    _style(ws3, 9)
    for col, w in zip("ABCDEFGHI", [6, 34, 6, 8, 6, 10, 10, 30, 26]):
        ws3.column_dimensions[col].width = w

    # 4) 성과 년도별 정리 (연장평가·실적)
    ws_perf = wb.create_sheet("성과 년도별 정리")
    ws_perf.append(["ID", "지표", "단위", "근거서류", "제출", "집계값", "근거(파일·페이지)", "상태", "비고"])
    for p in judgment.get("performance", []):
        ws_perf.append([p["id"], p["지표"], p["단위"], p["근거서류"], p["제출"],
                        p["집계값"], p.get("근거(파일·페이지)", ""), p["상태"], p["비고"]])
    _style(ws_perf, 9)
    for r in range(2, ws_perf.max_row + 1):
        st = ws_perf.cell(r, 8).value
        ws_perf.cell(r, 8).fill = PatternFill("solid", fgColor=COLORS.get(st, "FFFFFF"))
    for col, w in zip("ABCDEFGHI", [6, 16, 6, 24, 6, 16, 28, 10, 30]):
        ws_perf.column_dimensions[col].width = w

    # 4b) 재무 항목(연도별) — 재무제표에서 추출한 셀 + 출처
    fin = judgment.get("financials") or {}
    if fin.get("items"):
        src = (fin.get("file") or "") + (f" p.{fin['page']}" if fin.get("page") else "")
        wsf = wb.create_sheet("재무 항목")
        wsf.append(["항목", "연도", "값", "출처(파일·페이지)"])
        for item, yv in fin["items"].items():
            for yr, val in (yv or {}).items():
                wsf.append([item, yr, val, src])
        _style(wsf, 4)
        for col, w in zip("ABCD", [18, 12, 18, 28]):
            wsf.column_dimensions[col].width = w

    # 5) 서류 처리 로그
    ws4 = wb.create_sheet("서류 처리내역")
    ws4.append(["파일", "분류 유형", "신뢰도", "추출방식", "글자수", "필드수", "불러옴"])
    for d in record.get("doc_log", []):
        ws4.append([d["file"], d["유형"], d["신뢰도"], d["추출방식"],
                    d.get("글자수", 0), d["필드수"], d.get("불러옴", "")])
    _style(ws4, 7)
    for col, w in zip("ABCDEFG", [40, 22, 8, 10, 8, 8, 8]):
        ws4.column_dimensions[col].width = w

    wb.save(out_path)
    return out_path
